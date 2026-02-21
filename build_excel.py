"""
ClearMetric Cost of Living Comparator — Premium Excel Template
Product T10 for Gumroad ($11.99)

3 Sheets:
  1. City Comparator — pick 2 cities (dropdown), see equivalent salary + category breakdown
  2. Multi-City Ranking — all 30 cities ranked by overall index, housing, groceries, etc.
  3. How To Use — instructions

Design: Teal/Cyan palette (#0097A7 primary, #00838F dark, #E0F7FA input)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os

# Import city data
from data import CITY_DATA, CITIES, CATEGORIES

# ============================================================
# DESIGN SYSTEM — Teal/Cyan
# ============================================================
PRIMARY = "0097A7"   # Teal
DARK = "00838F"     # Dark teal
WHITE = "FFFFFF"
INPUT_TINT = "E0F7FA"  # Light cyan
LIGHT_BG = "F5FAFA"
MED_GRAY = "5D6D7E"
GREEN = "27AE60"
LIGHT_GREEN = "EAFAF1"

FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=12, color="B2EBF2", italic=True)
FONT_SECTION = Font(name="Calibri", size=13, bold=True, color=WHITE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_LABEL = Font(name="Calibri", size=11, color=DARK)
FONT_INPUT = Font(name="Calibri", size=12, color=DARK, bold=True)
FONT_VALUE = Font(name="Calibri", size=11, color=DARK)
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=DARK)
FONT_SMALL = Font(name="Calibri", size=9, color=MED_GRAY, italic=True)

FILL_PRIMARY = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid")
FILL_DARK = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
FILL_INPUT = PatternFill(start_color=INPUT_TINT, end_color=INPUT_TINT, fill_type="solid")
FILL_GRAY = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

THIN = Border(
    left=Side("thin", MED_GRAY),
    right=Side("thin", MED_GRAY),
    top=Side("thin", MED_GRAY),
    bottom=Side("thin", MED_GRAY),
)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center")


def cols(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


def header_bar(ws, row, c1, c2, text, fill=None):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=text)
    cell.font = FONT_SECTION
    cell.fill = fill or FILL_PRIMARY
    cell.alignment = ALIGN_C
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = fill or FILL_PRIMARY
        ws.cell(row=row, column=c).border = THIN


# Data sheet name for formulas (Excel requires quotes for names with spaces)
DATA_SHEET = "'Multi-City Ranking'"

# ============================================================
# SHEET 1: CITY COMPARATOR
# ============================================================
def build_city_comparator(ws):
    ws.title = "City Comparator"
    ws.sheet_properties.tabColor = PRIMARY
    cols(ws, {"A": 2, "B": 22, "C": 14, "D": 14, "E": 14})

    for r in range(1, 50):
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    # Title
    for r in range(1, 4):
        for c in range(2, 6):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:E1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:E2")
    ws.row_dimensions[2].height = 38
    title = ws.cell(row=2, column=2, value="CITY COMPARATOR")
    title.font = FONT_TITLE
    title.alignment = ALIGN_C
    ws.merge_cells("B3:E3")
    ws.row_dimensions[3].height = 22
    sub = ws.cell(row=3, column=2, value="Pick 2 cities. Enter your salary. See equivalent salary and category breakdown.")
    sub.font = FONT_SUBTITLE
    sub.alignment = ALIGN_C

    # Inputs
    ws.cell(row=5, column=2, value="Current City").font = FONT_LABEL
    ws.cell(row=5, column=2).fill = FILL_GRAY
    ws.cell(row=5, column=2).border = THIN
    ws.cell(row=5, column=2).alignment = ALIGN_L
    ws.cell(row=5, column=3, value="New York").font = FONT_INPUT
    ws.cell(row=5, column=3).fill = FILL_INPUT
    ws.cell(row=5, column=3).border = THIN
    ws.cell(row=5, column=3).alignment = ALIGN_L

    ws.cell(row=6, column=2, value="Target City").font = FONT_LABEL
    ws.cell(row=6, column=2).fill = FILL_GRAY
    ws.cell(row=6, column=2).border = THIN
    ws.cell(row=6, column=2).alignment = ALIGN_L
    ws.cell(row=6, column=3, value="Austin").font = FONT_INPUT
    ws.cell(row=6, column=3).fill = FILL_INPUT
    ws.cell(row=6, column=3).border = THIN
    ws.cell(row=6, column=3).alignment = ALIGN_L

    ws.cell(row=7, column=2, value="Your Annual Salary ($)").font = FONT_LABEL
    ws.cell(row=7, column=2).fill = FILL_GRAY
    ws.cell(row=7, column=2).border = THIN
    ws.cell(row=7, column=2).alignment = ALIGN_L
    ws.cell(row=7, column=3, value=100000).font = FONT_INPUT
    ws.cell(row=7, column=3).fill = FILL_INPUT
    ws.cell(row=7, column=3).number_format = "$#,##0"
    ws.cell(row=7, column=3).border = THIN
    ws.cell(row=7, column=3).alignment = ALIGN_R

    # Data validation for city dropdowns (reference Data sheet)
    dv1 = DataValidation(type="list", formula1=f"{DATA_SHEET}!$A$2:$A$31", allow_blank=False)
    dv1.error = "Pick a city from the list"
    ws.add_data_validation(dv1)
    dv1.add("C5")
    dv1.add("C6")

    # Results section
    header_bar(ws, 9, 2, 5, "Results")
    ws.cell(row=10, column=2, value="Equivalent Salary in Target City").font = FONT_LABEL
    ws.cell(row=10, column=2).fill = FILL_GRAY
    ws.cell(row=10, column=2).border = THIN
    ws.cell(row=10, column=3, value=f"=C7*INDEX({DATA_SHEET}!$B$2:$B$31,MATCH(C6,{DATA_SHEET}!$A$2:$A$31,0))/INDEX({DATA_SHEET}!$B$2:$B$31,MATCH(C5,{DATA_SHEET}!$A$2:$A$31,0))").font = FONT_BOLD
    ws.cell(row=10, column=3).number_format = "$#,##0"
    ws.cell(row=10, column=3).fill = FILL_WHITE
    ws.cell(row=10, column=3).border = THIN
    ws.cell(row=10, column=3).alignment = ALIGN_R

    # Category breakdown
    header_bar(ws, 12, 2, 5, "Category Breakdown (NYC = 100)")
    headers = ["Category", "Current City", "Target City", "Difference"]
    for i, h in enumerate(headers):
        c = 2 + i
        cell = ws.cell(row=13, column=c, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_PRIMARY
        cell.border = THIN
        cell.alignment = ALIGN_C

    # Data layout: A=City, B=Overall, C=Housing, D=Groceries, E=Transportation, F=Healthcare, G=Utilities
    for i, cat in enumerate(CATEGORIES):
        r = 14 + i
        col_num = i + 2  # 2=Housing, 3=Groceries, etc. (B=1, C=2, ... in range B:G)
        ws.cell(row=r, column=2, value=cat.replace("_", " ").title()).font = FONT_LABEL
        ws.cell(row=r, column=2).fill = FILL_GRAY
        ws.cell(row=r, column=2).border = THIN
        ws.cell(row=r, column=2).alignment = ALIGN_L
        ws.cell(row=r, column=3, value=f'=INDEX({DATA_SHEET}!$B$2:$G$31,MATCH($C$5,{DATA_SHEET}!$A$2:$A$31,0),{col_num})').font = FONT_VALUE
        ws.cell(row=r, column=3).number_format = "0"
        ws.cell(row=r, column=3).fill = FILL_WHITE
        ws.cell(row=r, column=3).border = THIN
        ws.cell(row=r, column=3).alignment = ALIGN_R
        ws.cell(row=r, column=4, value=f'=INDEX({DATA_SHEET}!$B$2:$G$31,MATCH($C$6,{DATA_SHEET}!$A$2:$A$31,0),{col_num})').font = FONT_VALUE
        ws.cell(row=r, column=4).number_format = "0"
        ws.cell(row=r, column=4).fill = FILL_WHITE
        ws.cell(row=r, column=4).border = THIN
        ws.cell(row=r, column=4).alignment = ALIGN_R
        ws.cell(row=r, column=5, value=f"=D{r}-C{r}").font = FONT_BOLD
        ws.cell(row=r, column=5).number_format = "+0;-0;0"
        ws.cell(row=r, column=5).fill = FILL_WHITE
        ws.cell(row=r, column=5).border = THIN
        ws.cell(row=r, column=5).alignment = ALIGN_R

    ws.protection.sheet = True
    for r in [5, 6, 7]:
        ws.cell(row=r, column=3).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 2: MULTI-CITY RANKING (data source + rankings)
# ============================================================
def build_multi_city(wb):
    ws = wb.create_sheet("Multi-City Ranking")
    ws.sheet_properties.tabColor = "00838F"
    cols(ws, {"A": 18, "B": 10, "C": 10, "D": 10, "E": 12, "F": 10, "G": 10, "H": 12})

    # Headers
    headers = ["City", "Overall", "Housing", "Groceries", "Transportation", "Healthcare", "Utilities", "Rank (Overall)"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=1, column=i + 1, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_DARK
        cell.border = THIN
        cell.alignment = ALIGN_C

    # Data rows (sorted by city name)
    for i, city in enumerate(CITIES):
        d = CITY_DATA[city]
        r = 2 + i
        ws.cell(row=r, column=1, value=city).font = FONT_LABEL
        ws.cell(row=r, column=1).border = THIN
        ws.cell(row=r, column=2, value=d["overall"]).font = FONT_VALUE
        ws.cell(row=r, column=2).border = THIN
        ws.cell(row=r, column=2).alignment = ALIGN_R
        for j, cat in enumerate(CATEGORIES):
            ws.cell(row=r, column=3 + j, value=d[cat]).font = FONT_VALUE
            ws.cell(row=r, column=3 + j).border = THIN
            ws.cell(row=r, column=3 + j).alignment = ALIGN_R
        ws.cell(row=r, column=8, value=f"=RANK(B{r},$B$2:$B$31,1)").font = FONT_VALUE
        ws.cell(row=r, column=8).border = THIN
        ws.cell(row=r, column=8).alignment = ALIGN_R


# ============================================================
# SHEET 3: HOW TO USE
# ============================================================
def build_instructions(wb):
    ws = wb.create_sheet("How To Use")
    ws.sheet_properties.tabColor = MED_GRAY
    cols(ws, {"A": 3, "B": 90})

    ws.merge_cells("A1:B2")
    c = ws.cell(row=1, column=1, value="HOW TO USE THE COST OF LIVING COMPARATOR")
    c.font = FONT_TITLE
    c.fill = FILL_DARK
    c.alignment = ALIGN_C
    for r in range(1, 3):
        for co in range(1, 3):
            ws.cell(row=r, column=co).fill = FILL_DARK

    sections = [
        ("QUICK START", [
            "1. Open the 'City Comparator' tab",
            "2. Select your Current City and Target City from the dropdowns",
            "3. Enter your current annual salary in the input cell",
            "4. See your equivalent salary in the target city and category-by-category breakdown",
        ]),
        ("CITY COMPARATOR", [
            "Current City: Where you live now (or your salary's city)",
            "Target City: Where you're considering moving",
            "Your Annual Salary: Your current gross annual salary",
            "Equivalent Salary: The salary you'd need in the target city to maintain the same purchasing power",
            "Category Breakdown: Housing, groceries, transportation, healthcare, utilities — each indexed (NYC = 100)",
        ]),
        ("MULTI-CITY RANKING", [
            "See all 30 cities ranked by overall cost of living index",
            "Lower index = more affordable. NYC = 100 as baseline.",
            "Use this to compare multiple cities at a glance",
        ]),
        ("INDEX EXPLAINED", [
            "All indices use New York City as 100. A city with 80 is ~20% cheaper overall.",
            "Housing typically varies the most between cities.",
            "Groceries and utilities tend to be more similar across the US.",
        ]),
        ("IMPORTANT NOTES", [
            "This template is for educational purposes only. Not financial advice.",
            "Indices are estimates based on typical costs. Your actual costs may vary.",
            "© 2026 ClearMetric. For personal use only.",
        ]),
    ]

    r = 4
    for title, items in sections:
        ws.cell(row=r, column=2, value=title).font = Font(name="Calibri", size=12, bold=True, color=DARK)
        ws.cell(row=r, column=2).fill = PatternFill(start_color=INPUT_TINT, end_color=INPUT_TINT, fill_type="solid")
        ws.cell(row=r, column=2).border = THIN
        r += 1
        for item in items:
            ws.cell(row=r, column=2, value=item).font = Font(name="Calibri", size=11, color=DARK)
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 22
            r += 1
        r += 1


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default Sheet

    # Build Multi-City Ranking FIRST (City Comparator references it)
    build_multi_city(wb)

    # Build City Comparator (references 'Multi-City Ranking')
    ws_comp = wb.create_sheet("City Comparator", 0)
    build_city_comparator(ws_comp)

    # Build How To Use
    build_instructions(wb)

    # Set City Comparator as active sheet
    wb.active = wb["City Comparator"]

    out = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "output",
        "ClearMetric-Cost-of-Living-Comparator.xlsx",
    )
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Size: {os.path.getsize(out) / 1024:.1f} KB")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()